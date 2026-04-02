import csv
from datetime import datetime
from typing import List, Dict, Any, Optional
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

import config

_VN_TZ = ZoneInfo("Asia/Ho_Chi_Minh")
_UTC_TZ = ZoneInfo("UTC")


def _format_scan_time_vn(value: Any) -> str:
    """Hiển thị thời điểm quét URL; chuỗi SQLite (CURRENT_TIMESTAMP = UTC) → giờ VN."""
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    try:
        dt = datetime.fromisoformat(s.replace(" ", "T", 1))
    except ValueError:
        return s
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=_UTC_TZ)
    return dt.astimezone(_VN_TZ).strftime("%d/%m/%Y %H:%M:%S")


# Delta highlight (Sold T2 - Sold T1): red if > 5, yellow if > 3 (red takes precedence)
_FILL_DELTA_RED = PatternFill(start_color="FFE8B4B4", end_color="FFE8B4B4", fill_type="solid")
_FILL_DELTA_YELLOW = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")


def _fill_for_delta(delta: Any) -> Optional[PatternFill]:
    if delta is None:
        return None
    try:
        d = int(delta)
    except (TypeError, ValueError):
        return None
    if d > 5:
        return _FILL_DELTA_RED
    if d > 3:
        return _FILL_DELTA_YELLOW
    return None

def export_to_csv(analysis_results: List[Dict[str, Any]], 
                  keywords_results: List[Dict[str, Any]],
                  session_id: int) -> tuple:
    """
    Export analysis and keywords to CSV files
    
    Args:
        analysis_results: Product analysis data
        keywords_results: Keyword extraction data
        session_id: Session ID for filename
    
    Returns:
        Tuple of (products_csv_path, keywords_csv_path)
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Products CSV
    products_csv = config.OUTPUT_DIR / f"products_session_{session_id}_{timestamp}.csv"
    
    with open(products_csv, 'w', newline='', encoding='utf-8') as f:
        if analysis_results:
            fieldnames = [
                'rank_by_growth',
                'product_url',
                'product_title',
                'scanned_at_t1',
                'scanned_at_t2',
                'sold_t1',
                'sold_t2',
                'sold_delta',
                'growth_rate_percent',
            ]
            
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for row in analysis_results:
                rg = row.get('rank_by_growth')
                writer.writerow({
                    'rank_by_growth': rg if rg is not None else 'N/A',
                    'product_url': row['product_url'],
                    'product_title': row['product_title'],
                    'scanned_at_t1': _format_scan_time_vn(row.get('scanned_at_t1')),
                    'scanned_at_t2': _format_scan_time_vn(row.get('scanned_at_t2')),
                    'sold_t1': row['sold_t1'],
                    'sold_t2': row['sold_t2'],
                    'sold_delta': row['delta'],
                    'growth_rate_percent': f"{row['growth_rate']:.2%}",
                })
    
    # Keywords CSV
    keywords_csv = config.OUTPUT_DIR / f"keywords_session_{session_id}_{timestamp}.csv"
    
    with open(keywords_csv, 'w', newline='', encoding='utf-8') as f:
        if keywords_results:
            fieldnames = [
                'rank',
                'keyword_seo',
                'keyword_niche',
                'keyword_type',
                'frequency',
                'keyword_win_check',
            ]

            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()

            for row in keywords_results:
                writer.writerow({
                    'rank': row['rank'],
                    'keyword_seo': row.get('keyword_seo', ''),
                    'keyword_niche': row.get('keyword_niche', ''),
                    'keyword_type': row['keyword_type'],
                    'frequency': row['frequency'],
                    'keyword_win_check': row.get('keyword_win_check', ''),
                })
    
    return str(products_csv), str(keywords_csv)

def export_to_excel(analysis_results: List[Dict[str, Any]],
                    keywords_results: List[Dict[str, Any]],
                    session_id: int) -> str:
    """
    Export analysis and keywords to an easy-to-read Excel workbook.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = config.OUTPUT_DIR / f"report_session_{session_id}_{timestamp}.xlsx"

    wb = Workbook()
    ws_products = wb.active
    ws_products.title = "Products"

    product_headers = [
        "Rank by Growth",
        "Product URL",
        "Product Title",
        "Thời gian quét T1",
        "Thời gian quét T2",
        "Sold T1",
        "Sold T2",
        "Sold Delta",
        "Growth Rate (%)",
    ]
    ws_products.append(product_headers)

    for row in analysis_results:
        rg = row.get("rank_by_growth")
        ws_products.append([
            rg if rg is not None else "N/A",
            row["product_url"],
            row["product_title"],
            _format_scan_time_vn(row.get("scanned_at_t1")),
            _format_scan_time_vn(row.get("scanned_at_t2")),
            row["sold_t1"],
            row["sold_t2"],
            row["delta"],
            f"{row['growth_rate'] * 100:.2f}%",
        ])

    ncols = len(product_headers)
    for r_idx, arow in enumerate(analysis_results, start=2):
        fill = _fill_for_delta(arow.get("delta"))
        if fill:
            for c in range(1, ncols + 1):
                ws_products.cell(row=r_idx, column=c).fill = fill

    ws_keywords = wb.create_sheet("Keywords")
    keyword_headers = [
        "Rank",
        "Keyword tối ưu SEO",
        "Keyword niche",
        "Loại",
        "Tần suất",
        "Keyword cần kiểm tra mức độ win",
    ]
    ws_keywords.append(keyword_headers)

    for kw in keywords_results:
        ws_keywords.append([
            kw["rank"],
            kw.get("keyword_seo", ""),
            kw.get("keyword_niche", ""),
            kw["keyword_type"],
            kw["frequency"],
            kw.get("keyword_win_check", ""),
        ])

    for ws in [ws_products, ws_keywords]:
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        for idx, _ in enumerate(ws[1], start=1):
            cell = ws.cell(row=1, column=idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill

        ws.freeze_panes = "A2"

        for col in ws.columns:
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
            width = min(max(max_len + 2, 12), 60)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = width

    wb.save(excel_path)
    return str(excel_path)


def export_snapshot_to_excel(snapshot_rows: List[Dict[str, Any]],
                             session_id: int,
                             snapshot_order: int) -> str:
    """
    Export a single snapshot (t1 or t2) to an easy-to-read Excel workbook.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_path = config.OUTPUT_DIR / (
        f"snapshot_{snapshot_order}_session_{session_id}_{timestamp}.xlsx"
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Snapshot {snapshot_order}"

    headers = [
        "Product URL",
        "Product Title",
        "Sold Count",
        "Price",
        "Status",
        "Error Message",
        "Timestamp",
    ]
    ws.append(headers)

    for row in snapshot_rows:
        ws.append([
            row.get("product_url"),
            row.get("product_title"),
            row.get("sold_count"),
            row.get("price"),
            row.get("status"),
            row.get("error_message"),
            row.get("timestamp"),
        ])

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for idx, _ in enumerate(ws[1], start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        width = min(max(max_len + 2, 12), 70)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = width

    wb.save(excel_path)
    return str(excel_path)

def print_summary(analysis_results: List[Dict[str, Any]], 
                  keywords_results: List[Dict[str, Any]]):
    """Print summary to console"""
    print("\n" + "="*80)
    print("📊 ANALYSIS SUMMARY")
    print("="*80)
    
    print(f"\n🏆 TOP 10 PRODUCTS BY GROWTH (Delta):\n")
    
    ranked = [p for p in analysis_results if p.get('rank_by_growth') is not None]
    ranked.sort(key=lambda x: x['rank_by_growth'])
    for i, product in enumerate(ranked[:10], 1):
        print(f"{i}. {product['product_title'][:60]}")
        print(f"   Sold: {product['sold_t1']} → {product['sold_t2']} "
              f"(+{product['delta']}, {product['growth_rate']:.1%} growth)")
        print(f"   URL: {product['product_url']}")
        print()
    
    print("\n" + "-"*80)
    print(f"\n🔑 TOP 10 KEYWORDS (SEO vs niche):\n")

    for kw in keywords_results[:10]:
        seo = kw.get("keyword_seo") or ""
        niche = kw.get("keyword_niche") or ""
        label = f"SEO: {seo}" if seo else f"niche: {niche}"
        print(
            f"{kw['rank']}. {label} "
            f"({kw['keyword_type']}) — {kw['frequency']} occurrences"
        )
    
    print("\n" + "="*80)
